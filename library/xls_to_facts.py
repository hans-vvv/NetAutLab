#!/usr/bin/python

"""
     Copyright (c) 2016 World Wide Technology, Inc.
     All rights reserved.
     Revision history:
     7 Apr 2016  |  1.0 - initial release
"""

DOCUMENTATION = '''
---
module: xls_to_facts.py
author: Matt Mullen, World Wide Technology
version_added: "1.0"
short_description: Read an Excel .xlsx file and output Ansible facts
description:
    - Read the XLS file specified and output Ansible facts in the form of a list with each
      element in the list as a dictionary using the column header as the key and the contents
      of the cell as the value. A dictionary is created for each sheet,  in the format spreadsheet_SheetName.
 
requirements:
    - The openpyxl Python module must be installed on the Ansible host. This can be installed using pip:
      sudo pip install openpyxl  
options:
    src:
        description:
            - The name of the Excel spreadsheet
        required: true
    
'''

EXAMPLES = '''
    Running the module from the command line:
      ansible localhost -m xls_to_facts -a src="example.xlsx" -M ~/ansible/library
   localhost | SUCCESS => {
    "ansible_facts": {
        "spreadsheet_Sheet1": [
            {
                "Hostname": "Switch-1",
                "Mgmt_ip": "10.0.0.1"
            },
            {
                "Hostname": "Switch-2",
                "Mgmt_ip": "10.0.0.2"
            },
            {
                "Hostname": "Switch-3",
                "Mgmt_ip": "10.0.0.3"
            }
        ],
        "spreadsheet_Sheet2": [
            {
                "Description": "To Spine-1",
                "Interface": "Ethernet1/1",
                "Interface_IP": "192.168.100.1/30"
            },
            {
                "Description": "To Spine-2",
                "Interface": "Ethernet1/2",
                "Interface_IP": "192.168.100.5/30"
            }
        ]
    },
    "changed": false
    In a role configuration, given a group and host entry:
      [access_switch]
      10.0.0.1  ansible_connection=local ansible_ssh_user=ansible_local_user hostname=Switch-1
      #
      $ cat xls_to_facts.yml
      ---
      - name: Test Role to import facts from Excel
        hosts: access_switch
        roles:
          - {role: xls_to_facts, debug: on}
      $ ansible-playbook xls_to_facts.yml --ask-vault
'''
import openpyxl

# ---------------------------------------------------------------------------
# read_xls_dict
# ---------------------------------------------------------------------------

def read_xls_dict(input_file):
    "Read the XLS file and return as Ansible facts"
    result = {"ansible_facts":{}}
    spreadsheet = {}
    try:
        wb = openpyxl.load_workbook(input_file)
        for sheet in wb.get_sheet_names():
            ansible_sheet_name = 'spreadsheet_' + sheet
            spreadsheet[ansible_sheet_name] = []
            current_sheet = wb.get_sheet_by_name(sheet)
            dict_keys = []
            for c in range(1,current_sheet.max_column + 1):
                dict_keys.append(current_sheet.cell(row=1,column=c).value)
            for r in range (2,current_sheet.max_row + 1):
                temp_dict = {}
                for c in range(1,current_sheet.max_column + 1):
                    temp_dict[dict_keys[c-1]] = current_sheet.cell(row=r,column=c).value
                spreadsheet[ansible_sheet_name].append(temp_dict)
    except IOError:
        return (1, "IOError on input file:%s" % input_file)

    result["ansible_facts"] = spreadsheet
    return (0, result)

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

from ansible.module_utils.basic import *

def main():
    " "
    module = AnsibleModule(argument_spec = dict(
             src = dict(required=True)
             ),
             check_invalid_arguments=False,
             add_file_common_args=True)

    code, response = read_xls_dict(module.params["src"])
    if code == 1:
        module.fail_json(msg=response)
    else:
        module.exit_json(**response)

    return code

main()
#
